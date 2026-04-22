import os
import threading
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook, load_workbook

import models


BASE_DIR = os.path.join("backups", "excel")
_LOCK = threading.Lock()


def _slug(value: str, fallback: str = "genel") -> str:
    raw = (value or "").strip().lower()
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in raw).strip("_")
    return cleaned or fallback


def _project_info(db, project_id: int | None):
    if not project_id:
        year = str(datetime.now().year)
        return year, "proje_bilinmiyor", f"{year}_PROJE_BILINMIYOR_GENEL.xlsx", "genel"
    project = db.query(models.Project).filter(models.Project.id == int(project_id)).first()
    if not project:
        year = str(datetime.now().year)
        return year, f"proje_{project_id}", f"{year}_PROJE_{project_id}_KAYIT.xlsx", "genel"
    year = str(project.yil or datetime.now().year)
    project_code = _slug(project.kod or f"proje_{project.id}", f"proje_{project.id}")
    project_name = _slug(project.name or "is_emri", "is_emri")
    work_order_name = f"{year}_{project_code}_{project_name}.xlsx"
    return year, project_code, work_order_name, project.kod or str(project.id)


def _ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "is_emri_hareketleri"
    ws.append(["zaman", "islem", "kullanici", "rol", "kaynak", "detay"])
    wb.save(path)
    return wb


def _ensure_sheet_headers(wb, sheet_name: str, headers: Iterable[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(headers))
    if ws.max_row == 0:
        ws.append(list(headers))
    return ws


def append_project_event(
    db,
    project_id: int | None,
    sheet_name: str,
    headers: list[str],
    values: list[str],
):
    year, project_code, work_order_name, _ = _project_info(db, project_id)
    file_path = os.path.join(BASE_DIR, year, project_code, work_order_name)
    with _LOCK:
        wb = _ensure_workbook(file_path)
        ws = _ensure_sheet_headers(wb, sheet_name, headers)
        ws.append(values)
        wb.save(file_path)
    return file_path


def get_project_excel_path(db, project_id: int | None):
    year, project_code, work_order_name, _ = _project_info(db, project_id)
    return os.path.join(BASE_DIR, year, project_code, work_order_name)
